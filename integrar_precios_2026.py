"""Integra listas de precios mayoristas preservando su fecha real.

Coloque los archivos fuente en ``data/precios_2026/`` para usar la carpeta
interna, o pase otra carpeta como argumento. Ejemplos::

    python integrar_precios_2026.py
    python integrar_precios_2026.py "C:\\Users\\acer\\Oficina\\Info. para Serie Agricola"

El resultado principal se escribe en la raíz del proyecto como
``PRECIOS_MAYORISTAS_INTEGRADO.csv``. También se actualiza una copia legacy
con el nombre ``PRECIOS_MAYORISTAS_2026_INTEGRADO.csv`` para compatibilidad.

Esta base de precios se mantiene separada de la base de cantidades 2024/2025:
los precios no se cruzan con cantidades, ni se usan para calcular
elasticidades o relaciones precio-cantidad.
"""

from __future__ import annotations

import argparse
import csv
import contextlib
import io
import re
import shutil
import sys
import unicodedata
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable


PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = PROJECT_DIR / "data" / "precios_2026"
OUTPUT_PATH = PROJECT_DIR / "PRECIOS_MAYORISTAS_INTEGRADO.csv"
LEGACY_OUTPUT_PATH = PROJECT_DIR / "PRECIOS_MAYORISTAS_2026_INTEGRADO.csv"

MONTHS = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
MONTH_NAMES = {number: name.title() for name, number in MONTHS.items() if name != "setiembre"}
FIELDS = [
    "fecha", "año", "mes", "rubro", "especie", "variedad", "mercado", "procedencia",
    "localidad_corrientes", "envase", "kg_bulto", "total_kilos", "unidad", "precio_observado",
    "unidad_precio_observado", "precio_kg_estimado", "metodo_conversion_precio", "confianza_conversion_precio",
    "precio", "precio_min", "precio_max", "precio_promedio",
    "archivo_origen",
]
TABULAR_EXTENSIONS = {".csv", ".xlsx", ".xls"}
IGNORE_RE = re.compile(r"(?:^~\$|desktop\.ini$|registro|vfru|vhor|2024|2025)", re.I)


def _ascii(value: Any) -> str:
    return unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode().lower()


def _filename_text(filename: str) -> str:
    return _ascii(Path(filename).name).replace("_", " ").replace("-", " ")


def infer_rubro_from_filename(filename: str) -> str:
    text = _filename_text(filename)
    if "hortal" in text:
        return "Hortalizas"
    if "frut" in text or re.search(r"\b rf\d", text):
        return "Frutas"
    return ""


def infer_month_from_filename(filename: str) -> str:
    text = _filename_text(filename)
    for month in MONTHS:
        if re.search(rf"\b{month}\b", text) or month in text:
            return month.title()
    # Daily files use RF020126 / RH020126 (DDMMYY).
    match = re.search(r"[rhf](\d{2})(\d{2})(\d{2})", text.replace(" ", ""), re.I)
    if match:
        return MONTH_NAMES.get(int(match.group(2)), "")
    return ""


def infer_year_from_filename(filename: str) -> int | str:
    text = _filename_text(filename)
    if re.search(r"(?:2026|26)\b", text) or re.search(r"\d{6}", text):
        return 2026
    return ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    # Repair common mojibake without altering valid Spanish accents.
    for bad, good in (("Ã¡", "á"), ("Ã©", "é"), ("Ã­", "í"), ("Ã³", "ó"), ("Ãº", "ú"), ("Ã±", "ñ"), ("¥", "Ñ")):
        text = text.replace(bad, good)
    return text.title()


def is_invalid_spreadsheet_value(value: Any) -> bool:
    return _key(value) in {"ref", "div0", "value", "name", "na", "null", "none", "nan"}


def normalize_species(value: Any) -> str:
    text = normalize_text(value)
    return "" if is_invalid_spreadsheet_value(text) else text.replace("Prom. Esp.", "Promedio Especie")


def normalize_variety(value: Any) -> str:
    return normalize_text(value)


def normalize_market(value: Any) -> str:
    text = normalize_text(value)
    key = _key(text)
    if key.upper() in {"MCBA", "MERCADOCENTRALBSAS", "MERCADOCENTRALDEBUENOSAIRES"}:
        return "Mercado Central de Buenos Aires"
    return text


def normalize_location(value: Any) -> str:
    text = normalize_text(value)
    equivalents = {
        "Bs As": "Buenos Aires", "Bs. As.": "Buenos Aires", "Caba": "Ciudad Autónoma de Buenos Aires",
        "Entre Rios": "Entre Ríos", "E. Rios": "Entre Ríos", "Cordoba": "Córdoba", "R Negro": "Río Negro",
        "R. Negro": "Río Negro", "Ctes.": "Corrientes", "Ctes": "Corrientes", "M.D.Plat": "Mar del Plata",
        "Tucuman": "Tucumán", "Sgo. Est.": "Santiago del Estero", "S. Pedro": "San Pedro", "Peru": "Perú",
    }
    return equivalents.get(text, text)


def normalize_procedencia(value: Any) -> str:
    return normalize_location(value)


def infer_market_from_source(filename: str, sheet_name: str | None = None) -> str:
    """Infer the commercial market from the known price-list sources.

    Martin Micelli is the Corrientes market source; the monthly fruit and
    vegetable lists are known to come from MCBA. This inference is intentionally
    separate from ``procedencia``, which is the geographic origin of a product
    when the original table explicitly reports it.
    """
    text = _ascii(f"{filename} {sheet_name or ''}")
    if "martin micelli" in text:
        return "Mercado de Corrientes"
    has_category = any(token in text for token in ("frutas", "frutras", "fruta", "hortalizas", "hortaliza"))
    has_2026 = "2026" in text or re.search(r"(?:enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)[ _-]*26", text)
    has_month = any(month in text for month in MONTHS)
    if has_category and has_2026 and has_month:
        return "Mercado Central de Buenos Aires"
    return ""


def normalize_unit(value: Any) -> str:
    return normalize_text(value).replace("Por Kg", "$/Kg")


def normalize_price_unit(value: Any, envase: str = "", is_martin: bool = False) -> str:
    """Normalize the unit of the observed price without assuming $/kg."""
    text = normalize_text(value)
    key = _key(text)
    if key and ("kg" in key or "kilo" in key):
        return "$/kg"
    if key and any(token in key for token in ("bulto", "cajon", "caja", "bolsa", "jaula", "envase", "atado", "docena", "unidad", "planta")):
        return "$/bulto" if any(token in key for token in ("bulto", "cajon", "caja", "bolsa", "jaula", "envase")) else text.lower()
    if is_martin and envase:
        return "$/bulto"
    return text.lower() if text else "sin especificar"


def _parse_numeric_price(value: Any) -> float | None:
    """Parse a source number while retaining zero for audit purposes."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if value == value else None
    text = str(value).strip().replace("$", "").replace(" ", "")
    if not text or text.lower() in {"nan", "nat", "none", "null", "-", "–"}:
        return None
    text = re.sub(r"[^0-9,.-]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        decimal = "," if text.rfind(",") > text.rfind(".") else "."
        thousands = "." if decimal == "," else ","
        text = text.replace(thousands, "").replace(decimal, ".")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts[:-1]) + "." + parts[-1] if len(parts[-1]) <= 2 else "".join(parts)
    elif "." in text:
        left, right = text.rsplit(".", 1)
        text = left.replace(".", "") + ("." + right if len(right) <= 2 else right)
    try:
        return float(text)
    except ValueError:
        return None


def parse_argentine_price(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if value != 0 else None
    text = str(value).strip().replace("$", "").replace(" ", "")
    if not text or text in {"-", "–", "0", "0,00", "0.00"}:
        return None
    text = re.sub(r"[^0-9,.-]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        # The last separator is the decimal separator: supports both locales.
        decimal = "," if text.rfind(",") > text.rfind(".") else "."
        thousands = "." if decimal == "," else ","
        text = text.replace(thousands, "").replace(decimal, ".")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts[:-1]) + "." + parts[-1] if len(parts[-1]) <= 2 else "".join(parts)
    elif "." in text:
        # In Argentine notation 2.500 means 2500; a dot followed by one or
        # two digits is more likely a decimal point (1250.50).
        left, right = text.rsplit(".", 1)
        text = left.replace(".", "") + ("." + right if len(right) <= 2 else right)
    try:
        result = float(text)
        return result if result != 0 else None
    except ValueError:
        return None


def parse_date_strict(value: Any) -> str:
    """Return an ISO date only when the source value contains a real date.

    Supports Excel/pandas datetimes, DD-MM-YYYY, DD/MM/YYYY, ISO dates and
    Excel serial dates. A filename is deliberately never considered here.
    """
    if value is None or (isinstance(value, float) and value != value):
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime().strftime("%Y-%m-%d")
        except (AttributeError, ValueError, TypeError):
            pass
    if isinstance(value, (int, float)) and 1 <= float(value) <= 100000:
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            return ""
    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat", "none", "null"}:
        return ""
    text = text.split(" ", 1)[0]
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def derive_year_month_from_date(iso_date: str) -> tuple[int | str, str]:
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(iso_date or "")):
        return "", ""
    month = int(iso_date[5:7])
    return int(iso_date[:4]), MONTH_NAMES.get(month, "")


def normalize_date(value: Any, fallback_month: str | None = None, fallback_year: int | str | None = None) -> str:
    """Backward-compatible wrapper; fallbacks never invent a day."""
    return parse_date_strict(value)


def _key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", _ascii(value))


def read_price_file(path: Path) -> list[dict[str, Any]]:
    """Read CSV/XLSX/XLS and return rows as dictionaries.

    XLS support uses the optional ``xlrd`` package (``pip install xlrd``).
    A failure is raised to the caller so it can warn and continue.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = path.read_bytes()
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = raw.decode(encoding)
                sample = text[:4096]
                dialect = csv.Sniffer().sniff(sample, delimiters=",;")
                return list(csv.DictReader(io.StringIO(text), dialect=dialect))
            except (UnicodeDecodeError, csv.Error):
                continue
        raise ValueError("no se pudo detectar la codificación o el separador CSV")
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        if "martin micelli" in _ascii(path.name):
            print(f"Martin Micelli - hojas detectadas: {workbook.sheetnames}")
            preferred = next((name for name in workbook.sheetnames if _key(name) == "cargadedatos"), workbook.sheetnames[0])
            sheet = workbook[preferred]
            print(f"Martin Micelli - hoja analizada: {preferred}")
        values = list(sheet.values)
        workbook.close()
    elif suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("falta xlrd para leer .xls (instalar con: pip install xlrd)") from exc
        # Some legacy BIFF files emit informational messages on stderr.
        with contextlib.redirect_stderr(io.StringIO()):
            book = xlrd.open_workbook(file_contents=path.read_bytes(), on_demand=True)
        sheet = book.sheet_by_index(0)
        values = [sheet.row_values(i) for i in range(sheet.nrows)]
        book.release_resources()
    else:
        raise ValueError(f"extensión no soportada: {suffix}")
    if not values:
        return []
    headers = [str(v).strip() if v is not None else "" for v in values[0]]
    if "martin micelli" in _ascii(path.name):
        print(f"Martin Micelli - columnas detectadas: {headers}")
        print(f"Martin Micelli - filas leídas: {max(len(values) - 1, 0)}")
    return [dict(zip(headers, row)) for row in values[1:] if any(v not in (None, "") for v in row)]


def _find(row: dict[str, Any], names: Iterable[str]) -> Any:
    wanted = {_key(name) for name in names}
    for name, value in row.items():
        if _key(name) in wanted:
            return value
    # Source workbooks often use labels such as "Precio Público Mínimo".
    # Use a conservative substring match only after exact matching fails.
    for name, value in row.items():
        candidate = _key(name)
        if any(candidate.startswith(item) or item in candidate for item in wanted if len(item) >= 5):
            return value
    return ""


def _date_from_source(source: str) -> str:
    match = re.search(r"[rhf](\d{2})(\d{2})(\d{2,4})", Path(source).name, re.I)
    if not match:
        return ""
    year = int(match.group(3)); year += 2000 if year < 100 else 0
    return f"{year:04d}-{int(match.group(2)):02d}-{int(match.group(1)):02d}"


def _make_rows(raw_rows: list[dict[str, Any]], source: str, container: str) -> list[dict[str, Any]]:
    filename = f"{container} {source}"
    is_martin = "martin micelli" in _ascii(filename)
    inferred_market = infer_market_from_source(filename)
    rubro = infer_rubro_from_filename(filename)
    month = infer_month_from_filename(source) or infer_month_from_filename(container)
    year_fallback = infer_year_from_filename(source) or infer_year_from_filename(container) or ""
    price_candidates = ("precio_promedio", "precio promedio", "precio publico", "precio", "importe", "mopk", "precio_publico")
    if is_martin:
        print(f"Martin Micelli - columnas candidatas de precio: {price_candidates}")
    output = []
    parsed_dates = []
    valid_price_count = 0
    for raw in raw_rows:
        source_date = parse_date_strict(_find(raw, ("fecha", "date")))
        # Filename inference is allowed only when the table has no valid date.
        row_date = source_date or _date_from_source(source)
        row_year, row_month = derive_year_month_from_date(source_date)
        if not source_date:
            row_year, row_month = year_fallback, month
        row_rubro = rubro
        raw_type = normalize_text(_find(raw, ("rubro", "tipo", "serie")))
        if is_invalid_spreadsheet_value(raw_type):
            raw_type = ""
        if not row_rubro and raw_type:
            type_key = _key(raw_type).upper()
            row_rubro = "Hortalizas" if "HORTAL" in type_key else "Frutas" if "FRUT" in type_key else raw_type
        species = normalize_species(_find(raw, ("especie", "esp", "producto", "productos", "articulo")))
        if not species or _key(species) in {"total", "totales"}:
            continue
        avg_raw = _find(raw, price_candidates)
        minimum_raw = _find(raw, ("precio_min", "minimo", "mipk", "precio_publico_minimo"))
        maximum_raw = _find(raw, ("precio_max", "maximo", "mapk", "precio_publico_maximo"))
        avg = _parse_numeric_price(avg_raw)
        minimum = _parse_numeric_price(minimum_raw)
        maximum = _parse_numeric_price(maximum_raw)
        observed_candidates = [value for value in (avg, minimum, maximum) if value is not None and value > 0]
        observed = observed_candidates[0] if observed_candidates else next((value for value in (avg, minimum, maximum) if value is not None), None)
        if observed is not None and observed > 0:
            valid_price_count += 1
        if source_date:
            parsed_dates.append(source_date)
        if is_martin and not any(parse_argentine_price(_find(raw, candidates)) is not None for candidates in (price_candidates, ("precio_min", "minimo", "mipk", "precio_publico_minimo"), ("precio_max", "maximo", "mapk", "precio_publico_maximo"))):
            # Keep the row and its real date; do not manufacture a price.
            pass
        locality = _find(raw, ("localidad_corrientes", "localidad de corrientes"))
        envase = normalize_text(_find(raw, ("envase", "presentacion", "presentación", "bulto", "env")))
        kg_bulto = _parse_numeric_price(_find(raw, ("kg_bulto", "kg/bulto", "kg bulto", "peso bulto", "peso envase", "kg")))
        total_kilos = _parse_numeric_price(_find(raw, ("total_kilos", "total de kilos", "total kilos")))
        unit_source = _find(raw, ("unidad_precio_observado", "unidad precio", "unidad", "unit", "precio por", "precio/kg", "mopk", "mipk", "mapk"))
        observed_unit = normalize_price_unit(unit_source, envase, is_martin)
        unit_key = _key(observed_unit)
        source_price_keys = {_key(name) for name in raw if any(token in _key(name) for token in ("mopk", "mipk", "mapk", "precio/kg", "precioporkg"))}
        if source_price_keys:
            observed_unit = "$/kg"
            unit_key = "kg"
        explicit_per_kg = observed_unit == "$/kg" or "kg" in unit_key or "kilo" in unit_key
        comparable_unit = observed_unit == "$/kg" or observed_unit == "$/bulto"
        estimated_kg = None
        conversion_method = "sin conversión: unidad no comparable"
        conversion_confidence = "No comparable"
        if observed is None:
            conversion_method = "sin conversión: precio observado faltante"
            conversion_confidence = "Baja"
        elif observed <= 0:
            conversion_method = "sin conversión: precio observado cero o inválido"
            conversion_confidence = "Baja"
        elif explicit_per_kg:
            estimated_kg = observed
            conversion_method = "precio informado por kg"
            conversion_confidence = "Alta"
        elif comparable_unit and kg_bulto is not None and kg_bulto > 0:
            estimated_kg = observed / kg_bulto
            conversion_method = "precio observado dividido por kg_bulto"
            conversion_confidence = "Media"
        elif comparable_unit:
            conversion_method = "sin conversión: kg_bulto faltante"
        elif envase:
            conversion_method = "sin conversión: envase no expresado en kg"
        if is_martin:
            row_rubro_key = _key(raw_type).upper()
            if "HORTAL" in row_rubro_key:
                row_rubro = "Hortalizas"
            elif "FRUT" in row_rubro_key:
                row_rubro = "Frutas"
        row = {
            "fecha": row_date,
            "año": row_year,
            "mes": row_month,
            "rubro": row_rubro,
            "especie": species,
            # In MARTIN MICELLI, TIPO is the product category (rubro), not a variety.
            "variedad": normalize_variety(_find(raw, ("variedad", "var"))),
            "mercado": inferred_market or normalize_market(_find(raw, ("mercado", "market"))),
            "procedencia": normalize_procedencia(_find(raw, ("procedencia", "proc", "origen"))),
            "localidad_corrientes": normalize_location(locality),
            "envase": envase,
            "kg_bulto": kg_bulto,
            "total_kilos": total_kilos,
            "unidad": normalize_unit(_find(raw, ("unidad", "unit"))) or observed_unit,
            "precio_observado": observed,
            "unidad_precio_observado": observed_unit,
            "precio_kg_estimado": estimated_kg,
            "metodo_conversion_precio": conversion_method,
            "confianza_conversion_precio": conversion_confidence,
            "precio": observed,
            "precio_min": minimum,
            "precio_max": maximum,
            "precio_promedio": avg,
            "archivo_origen": f"{container}/{source}" if container else source,
        }
        output.append(row)
    if is_martin:
        print(f"Martin Micelli - fecha mínima parseada: {min(parsed_dates) if parsed_dates else 'n/d'}")
        print(f"Martin Micelli - fecha máxima parseada: {max(parsed_dates) if parsed_dates else 'n/d'}")
        print(f"Martin Micelli - años detectados desde fecha: {sorted({d[:4] for d in parsed_dates}) or []}")
        print(f"Martin Micelli - precios válidos: {valid_price_count}")
        if not valid_price_count:
            print("ADVERTENCIA: no se detectaron precios en MARTIN MICELLI; no se inventarán valores.", file=sys.stderr)
    return output


def _is_ignored(name: str) -> bool:
    return bool(IGNORE_RE.search(Path(name).name))


def _iter_tabular_entries(root: Path):
    for path in sorted(root.rglob("*")):
        if path.is_file() and path.suffix.lower() in TABULAR_EXTENSIONS and not _is_ignored(path.name):
            yield path, path.name, ""

    def from_zip(zip_path: Path, blob: bytes | None = None, container: str = ""):
        with zipfile.ZipFile(io.BytesIO(blob) if blob is not None else zip_path) as archive:
            for info in archive.infolist():
                if info.is_dir() or _is_ignored(info.filename):
                    continue
                name = Path(info.filename).name
                data = archive.read(info)
                if zipfile.is_zipfile(io.BytesIO(data)):
                    yield from from_zip(zip_path, data, f"{container}/{name}".strip("/"))
                elif Path(name).suffix.lower() in TABULAR_EXTENSIONS:
                    yield data, name, f"{container}/{zip_path.name}".strip("/")

    for path in sorted(root.rglob("*")):
        if path.is_file() and (path.suffix.lower() == ".zip" or zipfile.is_zipfile(path)) and not _is_ignored(path.name):
            yield from from_zip(path)


def _read_entry(entry: Any, name: str) -> list[dict[str, Any]]:
    if isinstance(entry, Path):
        return read_price_file(entry)
    suffix = Path(name).suffix.lower()
    # read_price_file uses paths to support openpyxl/xlrd; use a named temporary
    # file only for archive members and remove it immediately after reading.
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as handle:
        handle.write(entry); temp_path = Path(handle.name)
    try:
        return read_price_file(temp_path)
    finally:
        temp_path.unlink(missing_ok=True)


def main() -> int:
    parser = argparse.ArgumentParser(description="Integra precios mayoristas diarios 2026")
    parser.add_argument("input_dir", nargs="?", type=Path, default=DEFAULT_INPUT_DIR,
                        help="carpeta fuente; por defecto data/precios_2026/")
    args = parser.parse_args()
    input_dir = args.input_dir.expanduser()
    if not input_dir.exists():
        print(f"ERROR: no existe la carpeta fuente: {input_dir}", file=sys.stderr)
        return 1

    found = processed = errors = 0
    rows: list[dict[str, Any]] = []
    months: set[str] = set(); rubros: set[str] = set(); species: set[str] = set()
    for entry, name, container in _iter_tabular_entries(input_dir):
        found += 1
        try:
            parsed = _make_rows(_read_entry(entry, name), name, container)
            rows.extend(parsed); processed += 1
            months.update(row["mes"] for row in parsed if row["mes"])
            rubros.update(row["rubro"] for row in parsed if row["rubro"])
            species.update(row["especie"] for row in parsed if row["especie"])
        except Exception as exc:
            errors += 1
            print(f"ADVERTENCIA: no se pudo leer {name}: {exc}", file=sys.stderr)

    with OUTPUT_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
    shutil.copyfile(OUTPUT_PATH, LEGACY_OUTPUT_PATH)
    print(f"Archivos encontrados: {found}")
    print(f"Archivos procesados correctamente: {processed}")
    print(f"Archivos con error: {errors}")
    print(f"Filas integradas: {len(rows)}")
    years = sorted({row["año"] for row in rows if row["año"]})
    dated = sorted(row["fecha"] for row in rows if row["fecha"])
    print(f"Archivo principal generado: {OUTPUT_PATH}")
    print(f"Filas por año: { {year: sum(row['año'] == year for row in rows) for year in years} }")
    print(f"Fecha mínima: {dated[0] if dated else 'n/d'}")
    print(f"Fecha máxima: {dated[-1] if dated else 'n/d'}")
    martin_rows = [row for row in rows if "martin micelli" in _ascii(row["archivo_origen"])]
    martin_dates = sorted(row["fecha"] for row in martin_rows if row["fecha"])
    print(f"Filas provenientes de Martin Micelli: {len(martin_rows)}")
    print(f"Fecha mínima Martin Micelli: {martin_dates[0] if martin_dates else 'n/d'}")
    print(f"Fecha máxima Martin Micelli: {martin_dates[-1] if martin_dates else 'n/d'}")
    print(f"Años Martin Micelli: {sorted({row['año'] for row in martin_rows if row['año']})}")
    valid_prices = [row for row in rows if _parse_numeric_price(row["precio_observado"]) is not None and _parse_numeric_price(row["precio_observado"]) > 0]
    print(f"Registros con precio válido: {len(valid_prices)}")
    print(f"Registros sin precio válido: {len(rows) - len(valid_prices)}")
    estimated_prices = [row for row in rows if _parse_numeric_price(row["precio_kg_estimado"]) is not None and _parse_numeric_price(row["precio_kg_estimado"]) > 0]
    print(f"Registros con precio por kg estimado: {len(estimated_prices)}")
    print(f"Registros no comparables: {sum(row['confianza_conversion_precio'] == 'No comparable' for row in rows)}")
    print(f"Unidades de precio observadas: {sorted({row['unidad_precio_observado'] for row in rows if row['unidad_precio_observado']})}")
    print(f"Confianza de conversión: { {value: sum(row['confianza_conversion_precio'] == value for row in rows) for value in ('Alta', 'Media', 'Baja', 'No comparable')} }")
    martin_envases = sorted({row['envase'] for row in martin_rows if row['envase']})
    martin_kg = sorted({row['kg_bulto'] for row in martin_rows if row['kg_bulto'] is not None})
    martin_zero_kg = sum(row['kg_bulto'] == 0 for row in martin_rows)
    martin_extreme = sum(row['precio_kg_estimado'] is not None and (row['precio_kg_estimado'] > 100000 or row['precio_kg_estimado'] < 1) for row in martin_rows)
    print(f"Martin Micelli - envases únicos: {martin_envases[:30]}{' ...' if len(martin_envases) > 30 else ''}")
    print(f"Martin Micelli - kg_bulto únicos: {martin_kg[:30]}{' ...' if len(martin_kg) > 30 else ''}")
    print(f"Martin Micelli - casos kg_bulto igual a cero: {martin_zero_kg}")
    print(f"Martin Micelli - conversiones con precio por kg extremo: {martin_extreme}")
    markets = sorted({row["mercado"] for row in rows if row["mercado"]})
    procedencias = sorted({row["procedencia"] for row in rows if row["procedencia"]})
    print(f"Mercados detectados: {', '.join(markets) or '(ninguno)'}")
    print(f"Porcentaje de mercado informado: {100 * sum(bool(row['mercado']) for row in rows) / len(rows):.1f}%" if rows else "Porcentaje de mercado informado: 0.0%")
    print(f"Procedencias detectadas: {', '.join(procedencias) or '(ninguna)'}")
    print(f"Porcentaje de procedencia informada: {100 * sum(bool(row['procedencia']) for row in rows) / len(rows):.1f}%" if rows else "Porcentaje de procedencia informada: 0.0%")
    print(f"Meses detectados: {', '.join(sorted(months)) or '(ninguno)'}")
    print(f"Rubros detectados: {', '.join(sorted(rubros)) or '(ninguno)'}")
    print(f"Especies únicas: {len(species)}")
    print(f"CSV generado: {OUTPUT_PATH}")
    print(f"Copia legacy actualizada: {LEGACY_OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
