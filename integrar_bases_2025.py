import csv
import os
from collections import OrderedDict

import openpyxl


PROJECT = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.join(os.path.dirname(PROJECT), "Bases-Serie Agricola")
CURRENT_CSV = os.path.join(PROJECT, "REGISTRO 2025.csv")
OUTPUT_CSV = os.path.join(PROJECT, "REGISTRO 2025 INTEGRADO.csv")

MONTHS = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def text(value):
    return "" if value is None else str(value).strip().upper()


def normalize_origin(value):
    value = text(value)
    return {
        "BS. AS.": "BUENOS AIRES",
        "BS AS": "BUENOS AIRES",
        "CTES.": "CORRIENTES",
        "CTES": "CORRIENTES",
    }.get(value, value)


def parse_weight(value):
    return float((value or "0").replace(".", "").replace(",", "."))


def key(row):
    return tuple(row[field] for field in (
        "FECHA", "MERCADO", "SERIE", "ESPECIE", "VARIEDAD", "PROCEDENCIA", "PESO"
    ))


def append_monthly_rows(rows, workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    for sheet_name, series in (
        ("MEN FRUTAS", "FRUTAS"),
        ("MEN HORT", "HORTALIZAS"),
        ("Tomate", "HORTALIZAS"),
        ("Pimiento", "HORTALIZAS"),
    ):
        sheet = workbook[sheet_name]
        headers = [text(cell.value).lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        month_columns = {
            MONTHS[header]: column + 1
            for column, header in enumerate(headers)
            if header in MONTHS
        }
        for values in sheet.iter_rows(min_row=2, values_only=True):
            species = text(values[0])
            if not species or species.startswith("TOTAL"):
                continue
            variety = text(values[1]) or "SIN VARIEDAD"
            origin = normalize_origin(values[2])
            for month, column in month_columns.items():
                tons = values[column - 1]
                if not isinstance(tons, (int, float)) or tons <= 0:
                    continue
                rows.append({
                    "FECHA": f"01/{month:02d}/2025",
                    "MERCADO": "BSAS",
                    "SERIE": series,
                    "ESPECIE": species,
                    "VARIEDAD": variety,
                    "PROCEDENCIA": origin,
                    "MUNICIPIO": "",
                    "PESO": f"{tons * 1000:.3f}".replace(".", ","),
                })


def append_v_rows(rows, workbook_path, sheet_name, series, year, months):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        species = text(values[0])
        if not species or species == "TOTAL":
            continue
        if text(values[1]) == "LA ESPECIE":
            continue
        variety = text(values[1]) or "SIN VARIEDAD"
        origin = normalize_origin(values[2])
        for month, column in months:
            tons = values[column - 1]
            if not isinstance(tons, (int, float)) or tons <= 0:
                continue
            rows.append({
                "FECHA": f"01/{month:02d}/{year}",
                "MERCADO": "BSAS",
                "SERIE": series,
                "ESPECIE": species,
                "VARIEDAD": variety,
                "PROCEDENCIA": origin,
                "MUNICIPIO": "",
                "PESO": f"{tons * 1000:.3f}".replace(".", ","),
            })


def main():
    rows = []
    with open(CURRENT_CSV, encoding="utf-8-sig", newline="") as handle:
        rows.extend(csv.DictReader(handle, delimiter=";"))

    monthly_rows = []
    append_monthly_rows(monthly_rows, os.path.join(BASE_DIR, "2025 FRUTAS Y HORTALIZAS (mensual).xlsx"))
    late_rows = []
    append_v_rows(late_rows, os.path.join(BASE_DIR, "VFRU25_T.xlsx"), "VFRU25_T", "FRUTAS", 2025, ((11, 14), (12, 15)))
    append_v_rows(late_rows, os.path.join(BASE_DIR, "VHOR25_T.xlsx"), "VHOR25_T", "HORTALIZAS", 2025, ((11, 14), (12, 15)))
    append_v_rows(late_rows, os.path.join(BASE_DIR, "VFRU24_T.xlsx"), "VFRU24_T", "FRUTAS", 2024, tuple((month, month + 3) for month in range(1, 13)))
    append_v_rows(late_rows, os.path.join(BASE_DIR, "VHOR24_T.xlsx"), "VHOR24_T", "HORTALIZAS", 2024, tuple((month, month + 3) for month in range(1, 13)))

    # Preserve the detailed Corrientes register as-is: identical-looking rows
    # may still represent separate operations because there is no transaction ID.
    # The monthly workbook has priority for Jan-Oct. VFRU/VHOR supply only Nov-Dec.
    merged = OrderedDict((f"current-{index}", row) for index, row in enumerate(rows))
    new_keys = set()
    for row in monthly_rows:
        dedupe_key = key(row)
        if dedupe_key not in new_keys:
            merged[f"monthly-{len(new_keys)}"] = row
            new_keys.add(dedupe_key)
    for row in late_rows:
        dedupe_key = key(row)
        if dedupe_key not in new_keys:
            merged[f"late-{len(new_keys)}"] = row
            new_keys.add(dedupe_key)

    fields = ["FECHA", "MERCADO", "SERIE", "ESPECIE", "VARIEDAD", "PROCEDENCIA", "MUNICIPIO", "PESO"]
    with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter=";")
        writer.writeheader()
        writer.writerows(merged.values())
    print(f"current={len(rows)} monthly={len(monthly_rows)} late={len(late_rows)} integrated={len(merged)}")


if __name__ == "__main__":
    main()
